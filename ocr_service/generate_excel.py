#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_excel.py
=================
Membuat file Excel Lembur Harian berdasarkan data hasil OCR,
mengikuti format template "2026 LEMBUR HARIAN PT G2B".

Usage:
    python generate_excel.py <json_data> <template_path> <output_path> <bulan_int> <tahun_int>

Contoh:
    python generate_excel.py '[{"nama":"NASAR","records":[...]}]' \
        /app/TEMPLATE/2026_LEMBUR_HARIAN.xlsx \
        /app/storage/exports/lembur_april_2026.xlsx \
        4 2026
"""

import sys
import json
import os
import shutil
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"error": "openpyxl tidak terinstall. Jalankan: pip install openpyxl"}))
    sys.exit(1)

# ─── Jadwal Shift (sama dengan ScheduleSeeder.php) ──────────────────────────
SHIFTS = [
    {'slug': 'SHIFT_1_WEEKDAY', 'day_type': 'weekday', 'time_in': '08:00', 'time_out': '17:00'},
    {'slug': 'SHIFT_2_WEEKDAY', 'day_type': 'weekday', 'time_in': '19:00', 'time_out': '03:00'},
    {'slug': 'SHIFT_1_WEEKEND', 'day_type': 'saturday', 'time_in': '08:00', 'time_out': '13:00'},
    {'slug': 'SHIFT_2_WEEKEND', 'day_type': 'saturday', 'time_in': '13:00', 'time_out': '17:00'},
    {'slug': 'LEMBUR_SHIFT_1',  'day_type': 'holiday',  'time_in': '08:00', 'time_out': '17:00'},
    {'slug': 'LEMBUR_SHIFT_2',  'day_type': 'holiday',  'time_in': '19:00', 'time_out': '03:00'},
]

HARI_WEEKDAY = {0: 'SENIN', 1: 'SELASA', 2: 'RABU', 3: 'KAMIS', 4: 'JUMAT', 5: 'SABTU', 6: 'MINGGU'}

BULAN_NAMES = {
    1: 'JANUARI', 2: 'FEBRUARI', 3: 'MARET', 4: 'APRIL',
    5: 'MEI', 6: 'JUNI', 7: 'JULI', 8: 'AGUSTUS',
    9: 'SEPTEMBER', 10: 'OKTOBER', 11: 'NOVEMBER', 12: 'DESEMBER'
}

# ─── Style helpers ───────────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style='thin'),   right=Side(style='thin'),
    top=Side(style='thin'),    bottom=Side(style='thin')
)
HEADER_FILL  = PatternFill(fill_type='solid', fgColor='4472C4')   # Biru template
TOTAL_FILL   = PatternFill(fill_type='solid', fgColor='D9E1F2')   # Biru muda
TITLE_FILL   = PatternFill(fill_type='solid', fgColor='4472C4')
MINGGU_FILL  = PatternFill(fill_type='solid', fgColor='FFE4E1')   # Pink muda (Minggu)
BOLD_WHITE   = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
BOLD_BLACK   = Font(name='Calibri', bold=True, size=10)
NORMAL_FONT  = Font(name='Calibri', size=9)
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT         = Alignment(horizontal='left',   vertical='center')


def parse_time(time_str, date_str=None, next_day=False):
    """Parse HH:MM:SS ke datetime pada tanggal tertentu."""
    if not time_str or time_str == '-':
        return None
    try:
        t = datetime.strptime(time_str.strip(), '%H:%M:%S')
        if date_str:
            base = datetime.strptime(date_str, '%Y-%m-%d')
            dt = base.replace(hour=t.hour, minute=t.minute, second=t.second)
            if next_day:
                dt += timedelta(days=1)
            return dt
        return t
    except ValueError:
        return None


def detect_shift(scan1_str, date_str):
    """Deteksi shift yang cocok berdasarkan tanggal dan jam scan in (±3 jam)."""
    if not scan1_str:
        return None
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return None

    weekday = d.weekday()  # 0=Senin, 6=Minggu
    is_saturday = weekday == 5
    is_sunday   = weekday == 6
    # Catatan: hari libur nasional tidak dideteksi di script ini (tidak ada API)
    # Anggap Minggu = holiday type
    if is_sunday:
        day_type = 'holiday'
    elif is_saturday:
        day_type = 'saturday'
    else:
        day_type = 'weekday'

    scan_t = parse_time(scan1_str)
    if not scan_t:
        return None
    scan_hour = scan_t.hour

    for shift in SHIFTS:
        if shift['day_type'] != day_type:
            continue
        shift_in_hour = int(shift['time_in'].split(':')[0])
        diff = abs(scan_hour - shift_in_hour)
        diff = min(diff, 24 - diff)
        if diff <= 3:
            return shift

    # Fallback
    return next((s for s in SHIFTS if s['day_type'] == day_type), None)


def calc_overtime(scan1_str, scan2_str, date_str):
    """
    Hitung lembur dalam jam.
    Returns: (normal_hours, double_hours, minggu)
    - normal_hours: lembur 0–3 jam
    - double_hours: lembur di atas 3 jam
    - minggu: 1 jika hari Minggu/holiday dan ada scan
    """
    if not scan2_str or scan2_str == '-':
        return 0, 0, 0
    if not scan1_str or scan1_str == '-':
        return 0, 0, 0

    try:
        d = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return 0, 0, 0

    weekday  = d.weekday()
    is_sunday = weekday == 6

    if is_sunday:
        return 0, 0, 1

    shift = detect_shift(scan1_str, date_str)
    if not shift:
        return 0, 0, 0

    scan_in_dt  = parse_time(scan1_str, date_str)
    scan_out_dt = parse_time(scan2_str, date_str)
    if not scan_in_dt or not scan_out_dt:
        return 0, 0, 0

    # Jam keluar shift
    sched_out_t = datetime.strptime(shift['time_out'], '%H:%M')
    sched_out   = d.replace(hour=sched_out_t.hour, minute=sched_out_t.minute, second=0)

    # Handle overnight shift (time_out < time_in → hari berikutnya)
    sched_in_t = datetime.strptime(shift['time_in'], '%H:%M')
    sched_in   = d.replace(hour=sched_in_t.hour, minute=sched_in_t.minute, second=0)
    if sched_out < sched_in:
        sched_out += timedelta(days=1)

    # Handle jika scan_out terlihat lebih awal dari sched_out (overnight)
    if scan_out_dt < sched_in:
        scan_out_dt += timedelta(days=1)

    total_seconds = (scan_out_dt - sched_out).total_seconds()
    TOLERANCE_SECONDS = 55 * 60  # 55 menit

    if total_seconds <= TOLERANCE_SECONDS:
        return 0, 0, 0

    total_hours = int(total_seconds // 3600)
    if total_hours <= 3:
        return total_hours, 0, 0
    else:
        return 3, total_hours - 3, 0


# ─── Excel writer ─────────────────────────────────────────────────────────────

def write_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    if number_format: cell.number_format = number_format
    return cell


def apply_row_style(ws, row, col_start, col_end, fill=None, font=None, border=None, alignment=None):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        if fill:      cell.fill      = fill
        if font:      cell.font      = font
        if border:    cell.border    = border
        if alignment: cell.alignment = alignment


def generate_excel(employees, template_path, output_path, bulan, tahun):
    """Generate Excel file mengikuti format template."""
    bulan_name = BULAN_NAMES.get(int(bulan), f'BULAN_{bulan}')

    # ── Cari sheet yang cocok di template ────────────────────────────────────
    # Template punya sheet 'APRIL', 'MARET', dll
    wb_template = openpyxl.load_workbook(template_path)
    sheet_names = wb_template.sheetnames

    # Cari sheet dengan nama bulan
    template_sheet = None
    for name in sheet_names:
        if bulan_name.upper() in name.upper().strip():
            template_sheet = wb_template[name]
            break

    # Buat workbook baru
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = bulan_name

    # ── Setup column widths ───────────────────────────────────────────────────
    col_widths = {
        1: 2,   # A (margin)
        2: 2,   # B (margin)
        3: 6,   # C = NIP/No
        4: 22,  # D = NAMA
        5: 10,  # E = HARI
        6: 14,  # F = TANGGAL
        7: 12,  # G = SCAN 1
        8: 12,  # H = SCAN 2
        9: 9,   # I = NORMAL
        10: 9,  # J = DOUBLE
        11: 9,  # K = MINGGU
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Title baris 1 ─────────────────────────────────────────────────────────
    ws.merge_cells('C1:K1')
    title_cell = ws['C1']
    title_cell.value = f"DATA SCANLOG LEMBUR HARIAN — {bulan_name} {tahun}"
    title_cell.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6  # spacer

    # ── Tulis data per karyawan ───────────────────────────────────────────────
    current_row = 3
    emp_no = 1

    for emp in employees:
        nama = emp.get('nama', '-')
        nip  = emp.get('nip', str(emp_no))
        records = emp.get('records', [])

        # Header row per karyawan
        ws.row_dimensions[current_row].height = 18
        ws.merge_cells(f'C{current_row}:K{current_row}')
        hdr_merge = ws.cell(row=current_row, column=3, value=f"No. {emp_no} — {nama}")
        hdr_merge.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        hdr_merge.fill = PatternFill(fill_type='solid', fgColor='4472C4')
        hdr_merge.alignment = LEFT
        current_row += 1

        # Kolom header
        col_headers = ['NIP', 'NAMA', 'HARI', 'TANGGAL', 'SCAN 1', 'SCAN 2', 'NORMAL', 'DOUBLE', 'MINGGU']
        cols = [3, 4, 5, 6, 7, 8, 9, 10, 11]
        ws.row_dimensions[current_row].height = 16
        for col, header in zip(cols, col_headers):
            write_cell(ws, current_row, col, header,
                       font=Font(name='Calibri', bold=True, size=9, color='FFFFFF'),
                       fill=PatternFill(fill_type='solid', fgColor='5B9BD5'),
                       alignment=CENTER,
                       border=THIN_BORDER)
        header_row = current_row
        current_row += 1

        # Baris data
        data_start = current_row
        total_normal = 0; total_double = 0; total_minggu = 0

        # Buat lookup dari records OCR
        records_by_date = {}
        for rec in records:
            records_by_date[rec['tanggal']] = rec

        # Loop semua hari dalam bulan (bukan hanya yang ada di PDF)
        import calendar
        days_in_month = calendar.monthrange(int(tahun), int(bulan))[1]

        for day in range(1, days_in_month + 1):
            date_str = f"{tahun}-{str(bulan).zfill(2)}-{str(day).zfill(2)}"
            try:
                d_obj = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                continue
            hari = HARI_WEEKDAY[d_obj.weekday()]

            rec = records_by_date.get(date_str, {})
            scan1 = rec.get('scan1') or ''
            scan2 = rec.get('scan2') or ''

            normal, double, minggu = calc_overtime(scan1, scan2, date_str) if (scan1 and scan2) else (0, 0, 0)
            # Hari Minggu tanpa data tetap Minggu jika ada scan
            if d_obj.weekday() == 6 and scan1:
                minggu = 1

            total_normal += normal
            total_double += double
            total_minggu += minggu

            is_sunday = d_obj.weekday() == 6
            row_fill = MINGGU_FILL if is_sunday else None

            ws.row_dimensions[current_row].height = 14

            # Tulis kolom: NIP, NAMA, HARI, TANGGAL, SCAN1, SCAN2, NORMAL, DOUBLE, MINGGU
            row_values = [
                nip, nama, hari,
                d_obj.strftime('%d/%m/%Y'),
                scan1, scan2,
                normal if normal else None,
                double if double else None,
                minggu if minggu else None,
            ]
            for col, val in zip(cols, row_values):
                cell = write_cell(ws, current_row, col, val,
                                  font=Font(name='Calibri', size=9,
                                            color='CC0000' if is_sunday else '000000'),
                                  fill=row_fill,
                                  alignment=CENTER if col != 4 else LEFT,
                                  border=THIN_BORDER)

            current_row += 1

        data_end = current_row - 1

        # Baris TOTAL
        ws.row_dimensions[current_row].height = 16
        write_cell(ws, current_row, 3, '', border=THIN_BORDER)
        write_cell(ws, current_row, 4, nama, font=BOLD_BLACK, fill=TOTAL_FILL, alignment=LEFT, border=THIN_BORDER)
        write_cell(ws, current_row, 5, '', border=THIN_BORDER)
        write_cell(ws, current_row, 6, '', border=THIN_BORDER)
        write_cell(ws, current_row, 7, 'TOTAL', font=BOLD_BLACK, fill=TOTAL_FILL, alignment=CENTER, border=THIN_BORDER)
        write_cell(ws, current_row, 8, '', border=THIN_BORDER)
        write_cell(ws, current_row, 9,  total_normal if total_normal else None,
                   font=BOLD_BLACK, fill=TOTAL_FILL, alignment=CENTER, border=THIN_BORDER)
        write_cell(ws, current_row, 10, total_double if total_double else None,
                   font=BOLD_BLACK, fill=TOTAL_FILL, alignment=CENTER, border=THIN_BORDER)
        write_cell(ws, current_row, 11, total_minggu if total_minggu else None,
                   font=BOLD_BLACK, fill=TOTAL_FILL, alignment=CENTER, border=THIN_BORDER)

        current_row += 1
        # Spacer
        current_row += 2

        emp_no += 1

    # ── Simpan file output ────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb_out.save(output_path)
    print(json.dumps({"success": True, "output": output_path}))


def main():
    if len(sys.argv) < 6:
        print(json.dumps({"error": "Usage: generate_excel.py <json_data> <template_path> <output_path> <bulan> <tahun>"}))
        sys.exit(1)

    json_data     = sys.argv[1]
    template_path = sys.argv[2]
    output_path   = sys.argv[3]
    bulan         = int(sys.argv[4])
    tahun         = int(sys.argv[5])

    try:
        employees = json.loads(json_data)
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"JSON tidak valid: {e}"}))
        sys.exit(1)

    generate_excel(employees, template_path, output_path, bulan, tahun)


if __name__ == '__main__':
    main()
